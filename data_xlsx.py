import csv
from openpyxl import Workbook
from datetime import datetime
import os


# Функція для підрахунку віку на основі дати народження
def calculate_age(birthdate):
    today = datetime.today()
    return today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))


# Функція для створення XLSX файлу з різними віковими категоріями
def create_xlsx_from_csv(csv_filename, xlsx_filename):
    try:
        # Перевіряємо наявність CSV-файлу
        if not os.path.exists(csv_filename):
            print("Помилка: CSV файл не знайдено!")
            return

        # Створюємо файл XLSX
        wb = Workbook()
        ws_all = wb.active
        ws_all.title = "all"

        ws_younger_18 = wb.create_sheet("younger_18")
        ws_18_45 = wb.create_sheet("18-45")
        ws_45_70 = wb.create_sheet("45-70")
        ws_older_70 = wb.create_sheet("older_70")

        # Заголовки для аркушів
        headers = ['№', 'Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік']
        for ws in [ws_all, ws_younger_18, ws_18_45, ws_45_70, ws_older_70]:
            ws.append(headers)

        # Читаємо дані з CSV файлу і заповнюємо аркуші
        with open(csv_filename, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Пропускаємо заголовок CSV

            row_num = 1
            for row in reader:
                birthdate = datetime.strptime(row[4], "%Y-%m-%d")
                age = calculate_age(birthdate)
                new_row = [row_num] + row[:3] + [row[4], age]
                ws_all.append(new_row)

                # Розподіляємо по віковим категоріям
                if age < 18:
                    ws_younger_18.append(new_row)
                elif 18 <= age <= 45:
                    ws_18_45.append(new_row)
                elif 45 < age <= 70:
                    ws_45_70.append(new_row)
                else:
                    ws_older_70.append(new_row)

                row_num += 1

        # Зберігаємо файл XLSX
        wb.save(xlsx_filename)
        print("Ok, файл XLSX успішно створено!")

    except PermissionError:
        print("Помилка: неможливо створити XLSX файл (ймовірно, він відкритий).")
    except Exception as e:
        print(f"Помилка: {e}")


# Запуск програми
if __name__ == "__main__":
    csv_file = 'employees.csv'
    xlsx_file = 'employees.xlsx'

    create_xlsx_from_csv(csv_file, xlsx_file)
